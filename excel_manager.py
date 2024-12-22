import openpyxl

class ExcelManager:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(filename)
        self.sheet = self.workbook.active

    def get_employees(self):
        employees = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            employees.append({
                "ID": row[0],
                "Name": row[1],
                "Position": row[2],
                "Schedule": row[3]
            })
        return employees

    def update_employee(self, emp_id, name, position, schedule):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == emp_id:
                row[1].value = name
                row[2].value = position
                row[3].value = schedule
                break
        self.workbook.save(self.filename)

    def export_to_excel(self, output_filename):
        self.workbook.save(output_filename)
